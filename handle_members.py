# pylint: disable=import-error
from typing import Any
import pandas as pd
import re
import numpy as np
import shutil
import os, sys
from pathlib import Path
from datetime import date
import time 
from time import strftime
import openpyxl
from openpyxl import Workbook
from openpyxl import load_workbook
from packages.encoder import base_encode, base_decode
from packages.utils import convert_countrycode, convert_postnr, \
    normalize_name, normalize_email, concat_special_cols, \
    normalize_postort, concat_group_id, add_comment_info, \
    validate_file, normalize_phonenumber

'''
Idea:
1. Automate login and export of all members (see next bullet)
2. Export All OL-member, including "målsman"
3. Use pre-created Excel-templates and fill with exported data
'''

# Update to correct timezone
os.environ["TZ"] = "Europe/Stockholm"
time.tzset()

today = date.today()
date_today = today.strftime("%Y-%m-%d")
timestamp = str(strftime("%Y-%m-%d_%H.%M")) # Timestamp to use for filenames

# Remeber start time
start_time = time.time()

# Config
path_in =  '/usr/src/app/files/contact-list/'             # Required base path
path_out = '/usr/src/app/files/contact-list/created/'     # Output path
youth_contactlist_template = '/usr/src/app/templates/template_youth_contactlist.xlsx'
contactlist_template = '/usr/src/app/templates/template_contactlist.xlsx'
# Rules to add to export - as information
_rules = set({})

# Groups of interest
youth_groups=['OL Grön 1', 'OL Grön 2', 'OL Vit-Gul', 'OL Orange-Violett', 'OL Junior'] # 'OL Ungdom vilande' intentionally left out
youth_coach_groups=['OL Ledare - Grön', 'OL Ledare - Vit-Gul', 'OL Ledare - Orange-Violett', 'OL Ledare - Junior']
other_groups=['OL Instruktörer', 'OL Tisdagsträning-sommar', 'OL Tisdagsträning-vinter', 'OL Wendelsbergsträning', 'OL Intagning och utsättning']

# Holding all groups
# all_groups = pd.DataFrame() # Empty

# Get args
if len(sys.argv) > 1:
    cmd  = sys.argv[1]

if len(sys.argv) > 2:
    io_export_file_name = sys.argv[2]
    validate_file(io_export_file_name, 2)

if len(sys.argv) > 3:
    output_file_name = sys.argv[3]
    print(output_file_name)
    #validate_file(output_file_name, 3)

# Functions

# def save_file_plain(file_name, df):
#     """
#     Save to Excel file
#     """
#     df.to_excel(file_name, index=False)
#     return df

def save_file(file_name, df, color = True):
    """
    Save to Excel file
    """
    # To get colors to work
    writer = pd.ExcelWriter(file_name, engine='xlsxwriter')
    df.to_excel(writer, index=False)

    writer.save()
    # df.to_excel(file_name, index=False)
    return df

def stats(text):
    """
    Utility function to print stats. Easy to disable...
    """
    if True:
        print(text)

def add_rule(rule):
    _rules.add(rule)

def get_rules():
    return _rules

def _read_io_file(file_name, columns = None):
    """
    Read from IO file and return dataframe. Converts incoming data.
    """
    #_dtype = {'Förnamn': 'string','Efternamn': 'string','Födelsedat./Personnr.': 'string', 'Medlemsnr.': 'string',
    #    'Telefon mobil': 'string', 'Telefon bostad': 'string', 'Telefon arbete': 'string', 'Hemtelefon': 'string', 
    #    'Mobiltelefon': 'string', 'Arbetstelefon': 'string', 'Övrig medlemsinfo': 'string'}
    _dtype = {'Förnamn': 'string','Efternamn': 'string','Födelsedat./Personnr.': 'string', 'Medlemsnr.': 'string',
        'Övrig medlemsinfo': 'string', 'Typ':'string', 'Grupp/Lag/Arbetsrum/Familj':'string'}
    _converters = {
        'Telefon mobil':normalize_phonenumber, 
        'Telefon bostad':normalize_phonenumber, 
        'Telefon arbete':normalize_phonenumber, 
        'Hemtelefon':normalize_phonenumber, 
        'Mobiltelefon':normalize_phonenumber, 
        'Arbetstelefon':normalize_phonenumber, 
        'E-post kontakt':normalize_email, 'E-post privat':normalize_email,
        'Kontakt 1 epost':normalize_email, 
        'Postnummer':convert_postnr, 
        'Kontaktadress - Postort':normalize_postort,
        'Folkbokföring - Postort':normalize_postort,
        'Postort':normalize_postort}
    if columns:
        return pd.read_excel(file_name, 
            usecols = columns,
            dtype = _dtype,
            converters = _converters) 
    else:
        return pd.read_excel(file_name,
            dtype = _dtype,
            converters = _converters)

def get_all_from_export(io_export_file_name):
    """
    Read in all columns of interest from exported persons
    """
    # Get IO Export
    raw_df = _read_io_file(io_export_file_name)
    raw_df.to_csv(path_out+'csv/_all.csv')
    stats("Antal medlemmar exporterade från IO: {} ({})".format(str(len(raw_df)), Path(io_export_file_name).name))

    # Convert to nice format
    output_df = pd.DataFrame()
    output_df['Förnamn'] = raw_df['Förnamn']
    output_df['Efternamn'] = raw_df['Efternamn']
    output_df['Parent'] = raw_df['Målsman']
    output_df['Födelsedatum'] = raw_df['Födelsedat./Personnr.']
    add_rule("'År i år' är inte exakt antal år, utan just 'år i år'")
    output_df['År i år'] = output_df['Födelsedatum'].apply(calculate_age_class)
    output_df['Mobil'] = raw_df['Telefon mobil']
    output_df['Telefon bostad'] = raw_df['Telefon bostad']
    add_rule("'E-post kontakt' används som 'E-post'")
    output_df['E-post'] = raw_df['E-post kontakt']
    add_rule("'Folkbokföringsadress' används som kontaktadress")
    output_df['Gatuadress'] = raw_df['Folkbokföring - Gatuadress']
    output_df['Postnummer'] = raw_df['Folkbokföring - Postnummer']
    output_df['Postort'] = raw_df['Folkbokföring - Postort']
    output_df['Grupp'] = raw_df['Grupp/Lag/Arbetsrum/Familj']
    output_df['UGrupp'] = raw_df['Grupp/Lag/Arbetsrum/Familj'].apply(only_youth_groups)
    output_df['Familj'] = raw_df['Familj']
    # Use encoded IdrottsID as id
    add_rule("Unikt ID per person har skapats")
    output_df['ID'] = raw_df['IdrottsID'].apply(lambda x: base_encode(int(x.replace('IID',''))))
    output_df['Registrerad'] = raw_df['Registreringsdatum']
    output_df['Prova-på'] = raw_df['Typ'].dropna().apply(lambda x: str(x).replace("P","Ja"))
    
    return output_df

def group_in_groups(groups, grp):
    '''
    Check if group 'A' is in list [' A','B','C'] (True)
    '''
    if groups is np.nan:
        return False

    l = list(groups.split(",")) 
    l =[x.strip() for x in l]
    return grp in l

def only_youth_groups(groups):
    """
    Filter out only Youth groups
    """
    result = []
    groups_str = str(groups)
    for grp in groups_str.split(','):
        if grp.strip() in youth_groups:
            result.append(grp.strip())
    
    if len(result) > 0: 
        return ",".join(result)
    else:
        return ""

def calculate_age_class(birth_date):
    '''
    Convert birth date of format 1931-01-21 to age for given current year (= not exact age)
    '''
    if pd.isna(birth_date):
        return ""
    return today.year - int(birth_date[0:4])

def names_to_key(fname,lname):
    '''
    Construct key to be able to map child with parent
    '''
    return (fname.strip() + '_' + lname.strip()).replace(' ','_').lower()

def parentinfo_to_key(info):
    '''
    Construct key to be able to map parent with child
    '''
    return info.replace('Till målsman för: ','').replace(' ','_').strip().lower()

def normalize_group_name(name, lowercase = False, strip_ol = True, ascii = False):
    '''
    Normalize group name to name to be used for file-names etc.
    '''
    name = name.strip().replace(' ','_')
    name = name.strip().replace('_-_','-')
    if strip_ol and name.startswith('OL_'):
        name = name.replace('OL_', '', 1)
    if lowercase:
        name = name.lower()
    if ascii:
        # Very simple conversion
        name = name.translate(name.maketrans("åäöÅÄÖ","aaoAAO"))
    return name

def save_templated_youth_excel(template, df, filename, rules):
    '''
    Saves Excel file for youth group
    - Opens Excel template
    - Adds data frame information
    - Save to new file for current data frame
    '''
    wb = Workbook()
    wb = load_workbook(template)

    # Set empty cells to ""
    df.fillna("", inplace=True)

    # Add rules info
    ws = wb["Information"]
    row = 15
    for rule in rules:
        ws.cell(column=2, row=row, value=rule)
        row += 1

    # Kontaktlista
    ws = wb["Kontaktlista"]
    for rowidx, row in df.iterrows():
        col = 1
        for c in row.values:
            #ws.cell(column=col, row=rowidx+2, value="{}".format(c))
            ws.cell(column=col, row=rowidx+1, value=c)
            col += 1

    # Närvarolista
    ws = wb["Närvarolista"]
    for rowidx, row in df[['Förnamn','Efternamn','År i år','ID']].iterrows():
        col = 1
        for c in row.values:
            ws.cell(column=col, row=rowidx+1, value=c)
            col += 1

    # Checklista
    ws = wb["Checklista"]
    for rowidx, row in df[['Förnamn','Efternamn','År i år','ID']].iterrows():
        col = 1
        for c in row.values:
            ws.cell(column=col, row=rowidx+1, value=c)
            col += 1

    wb.save(filename = filename)

def save_templated_excel(template, df, filename, rules):
    '''
    Saves Excel file for group
    - Opens Excel template
    - Adds data frame information
    - Save to new file for current data frame
    '''
    wb = Workbook()
    wb = load_workbook(template)

    # Set empty cells to ""
    df.fillna("", inplace=True)

    # Add rules info
    ws = wb["Information"]
    r = 15
    for rule in rules:
        ws.cell(column=2, row=r, value=rule)
        r += 1

    # Kontaktlista
    ws = wb["Kontaktlista"]
    for rowidx, row in df.iterrows():
        col = 1
        for c in row.values:
            ws.cell(column=col, row=rowidx+1, value=c)
            col += 1

    # Närvarolista
    ws = wb["Närvarolista"]
    for rowidx, row in df[['Förnamn','Efternamn','År i år','ID']].iterrows():
        col = 1
        for c in row.values:
            ws.cell(column=col, row=rowidx+1, value=c)
            col += 1

    # Checklista
    ws = wb["Checklista"]
    for rowidx, row in df[['Förnamn','Efternamn','År i år','ID']].iterrows():
        col = 1
        for c in row.values:
            ws.cell(column=col, row=rowidx+1, value=c)
            col += 1

    wb.save(filename = filename)

def calculate_type(birth_date):
    age = 0
    if pd.isna(birth_date):
        age = 0
    else:
        age = today.year - int(birth_date[0:4])
    return "U" if age < 25 else "V"

def create_youth_contactlist(name, df_src, df_p):
    '''
    Create contact list for given youth group
    - name: Name of this youth group
    - df_src: Data frame with only children of this group
    - df_p: Data frame with all parents (all groups)
    '''
    print("Create youth contactlist for group: {}".format(name))

    # Include persons only in current group
    df_c = df_src[df_src['UGrupp'].isin([name])].copy()

    # Merge parents with children - in two steps
    add_rule("Målsmän används som föräldrar för ungdomar")
    df_m = pd.merge(df_c, df_p[df_p['parent_no'] == 1], how='left', left_on='key', right_on='ref', suffixes=('','1'))
    df_m = pd.merge(df_m, df_p[df_p['parent_no'] == 2], how='left', left_on='key', right_on='ref', suffixes=('','2'))
    df_m.drop(columns=['Parent', 'Grupp', 'UGrupp', 'key', 'Parent1', 'Födelsedatum1', 'Registrerad1', 'Prova-på1',
        'År i år1', 'Grupp1', 'UGrupp1','Familj1', 'ref', 'parent_no',
        'Parent2', 'Födelsedatum2', 'Registrerad2', 'Prova-på2', 'År i år2', 
        'Grupp2', 'UGrupp2', 'Familj2', 'ref2', 'parent_no2'], inplace=True)
    
    add_rule("För ungdomsgrupper visas ungdomar först, vuxna sist")
    df_m.insert(4, 'Typ', df_m['Födelsedatum'].apply(calculate_type))

    # Sort
    df_m.sort_values(by=['Typ','Efternamn', 'Förnamn'], inplace=True)

    # Set index to 1..n
    df_m.index = np.arange(1,len(df_m)+1)

    # To Excel
    save_templated_youth_excel(youth_contactlist_template,df_m,path_out+normalize_group_name(name,True,False,True)+'_listor.xlsx', get_rules())
    
    # To CSV
    df_m.to_csv(path_out+'csv/'+normalize_group_name(name,True,False,True)+'_kontaktlista.csv')

    # To JSON - split seems to be best
    df_m[['Förnamn', 'Efternamn', 'ID', 'År i år']].to_json(path_out+'json/'+normalize_group_name(name,True,False,True)+'.json', orient='split', force_ascii=False)

    # Add to 'All groups'
    # all_groups.append(df_m)
    
def create_contactlist(name, df):
    '''
    Create contact list for given group name
    - name: Name of this group of interest
    - df: Group of persons 
    '''
    print("Create contactlist for group: {}".format(name))

    #df.drop(columns=['Parent', 'Grupp', 'UGrupp','Familj'], inplace=True)
    df.drop(columns=['Parent', 'Grupp', 'UGrupp'], inplace=True)
    # Set index to 1..n
    df.index = np.arange(1,len(df)+1)

    # To Excel
    save_templated_excel(contactlist_template,df,path_out+normalize_group_name(name,True,False,True)+'_listor.xlsx', get_rules())

    # To CSV
    df.sort_values(by=['Efternamn', 'Förnamn']).to_csv(path_out+'csv/'+normalize_group_name(name,True,False,True)+'_kontaktlista.csv')

    # To JSON
    df[['Förnamn', 'Efternamn', 'ID', 'År i år']].sort_values(by=['Efternamn', 'Förnamn']).to_json(path_out+'json/'+normalize_group_name(name,True,False,True)+'.json', orient='split', force_ascii=False)

    # Add to 'All groups'
    # all_groups.append(df)

def io_bug_detector(fname, lname, parent):
    '''
    Check if valid parent or not (=not same name as self)
    '''
    return True if (fname + " " + lname) ==  parent.replace("Till målsman för: ","") else False

# Action 
print(" Start ".center(80, "-"))

if "contact_list" == cmd:
    print("Create contact list file")
    df_all = get_all_from_export(io_export_file_name)
    #df_all.to_csv(path_out+'all.csv')

    # Get df for only parents
    df_parents = df_all[df_all['Parent'].notnull()].copy()

    # Idrott Online exports children without "målsmän" wrong
    # These children becomes parents ("målsmän") to themselves which must be wrong
    # Removing theses erroneous rows
    io_bug = df_parents.apply(lambda x: io_bug_detector(x['Förnamn'],x['Efternamn'],x['Parent']), axis=1)
    # Exclude persons that is parents to themselves - according to IO
    df_parents = df_parents.loc[~io_bug]

    # Get df for children (training in defined group)
    df_training_children = df_all[~df_all['Parent'].notnull() & df_all['UGrupp'].isin(youth_groups)].copy()
    
    # Construct key - so we can map this to parents later on
    df_training_children['key'] = df_training_children.apply(lambda x: names_to_key(x['Förnamn'],x['Efternamn']),axis=1)
 
    # Construct key ref - so we can map this to children later
    df_parents['ref'] = df_parents.apply(lambda x: parentinfo_to_key(x['Parent']),axis=1)
    df_parents['parent_no'] = df_parents.groupby(['ref'], dropna=False)['ref'].cumcount()+1

    # Create contact list for each youth group separately
    for grp in youth_groups:
        df = df_training_children[(df_training_children['Grupp'].apply(lambda x: group_in_groups(x,grp)))].copy()
        create_youth_contactlist(grp, df, df_parents)

    # Create contact lists for other groups separately 
    for grp in youth_coach_groups + other_groups:
        df = df_all[(~(df_all['Parent'].str.len()>0)) & (df_all['Grupp'].apply(lambda x: group_in_groups(x,grp)))].copy()
        create_contactlist(grp, df)

    #print("All: ")
    #print(all_groups)

def family_trick(el, df):
    return el if el != "" else f"((trick:{df.name}))"

def get_paper(famadmin, no_paper, in_familiy, in_household):
    """Evaluate of person should have a paper or not"""
    if famadmin == "Ja":
        return "Ja"
    elif no_paper != "": # Ingen tidning (MC)
        return "Nej"
    elif famadmin == "" and in_familiy == "Nej" and in_household == "Nej":
        return "Ja"
    return "Nej"

def get_my_club_id(comment):
    if pd.isna(comment):
        return ""
    regexp = r"\[\[MC-ID: (\d*)\]\]"
    regexp2 = r"\[\[MedlemsID: (\d*)\]\]"
    regexp3 = r"^Mc (\d*)$"
    regexp4 = r"^Mc(\d*)$"
    if re.search(regexp, comment):
        return re.search(regexp, comment).group(1) # [[MC-ID: nnnn]]
    elif re.search(regexp2, comment):
        return re.search(regexp2, comment).group(1) # [[Medlems-ID: nnn]]
    elif re.search(regexp3, comment):
        return re.search(regexp3, comment).group(1) # "Mc nnnn"
    elif re.search(regexp4, comment):
        return re.search(regexp4, comment).group(1) # "Mcnnnn"
    return ""

if "frisksport" == cmd:
    print("Create file for 'Frisksport'-paper")

    # Get members from file
    _dtype = {'Födelsedat./Personnr.': 'string', 'Medlemsnr.': 'string',
        'Telefon mobil': 'string', 'Telefon bostad': 'string', 'Telefon arbete': 'string', 'Hemtelefon': 'string', 
        'Mobiltelefon': 'string', 'Arbetstelefon': 'string', 'Övrig medlemsinfo': 'string'}
    # Must clean some data, due to poor quality (trailing spaces, not consistent case for families etc.)
    _converters = {'E-post kontakt':normalize_email, 'E-post privat':normalize_email,
        'Kontakt 1 epost':normalize_email,
        'Förnamn':normalize_name,
        'Efternamn':normalize_name,
        'Folkbokföring - Gatuadress':normalize_name,
        'Kontaktadress - Gatuadress':normalize_name,
        'Postnummer':convert_postnr, 
        'Kontaktadress - Postort':normalize_postort,
        'Postort':normalize_postort}
    df = pd.read_excel(io_export_file_name,
        dtype = _dtype,
        converters = _converters)

    df.rename(columns = {'Grupp/Lag/Arbetsrum/Familj':'Grupp', 'Födelsedat./Personnr.':'Personnummer'}, inplace = True)

    # Only use intersting columns
    interesting_cols = ['Prova-på','Förnamn','Efternamn','IdrottsID','Personnummer','Grupp','Familj','Fam.Admin',
        'Telefon mobil','Telefon bostad','Telefon arbete','E-post privat','E-post arbete',
        'E-post kontakt',
        'Folkbokföring - Gatuadress', 'Folkbokföring - Postnummer', 'Folkbokföring - Postort', 
        'Kontaktadress - Gatuadress', 'Kontaktadress - Postnummer', 'Kontaktadress - Postort', 'Kontaktadress - c/o adress', 'Kontaktadress - Land',
        #'Alt. förnamn','Kön','Nationalitet','Arbetsadress - Gatuadress','Arbetsadress - Postnummer','Arbetsadress - Postort','Arbetsadress - Land',
        'Arbetsadress - c/o adress',
        'Medlemsnr.','Medlem sedan','Medlem t.o.m.','Övrig medlemsinfo']
    df = pd.DataFrame(df, columns=interesting_cols).fillna("")

    # Keep one per family, keep 'Fam.Admin'
    df_ingen_tidning = df.sort_values(['Fam.Admin','Personnummer'])
    # Add value as trick when family="" (to keep all these for now)
    df_ingen_tidning['x'] = df_ingen_tidning.apply(lambda x: family_trick(x['Familj'],x), axis=1)

    # Ingen tidning
    df_ingen_tidning['Ingen tidning'] = np.where(df_ingen_tidning['Grupp'].str.contains('MC_IngenTidning'),"Ingen tidning (MC)","")

    # Del av familj
    df_ingen_tidning['Del av familj'] = np.where(df_ingen_tidning.duplicated(subset=['Familj','x'], keep=False),"Ja","Nej")

    # Del av hushåll
    df_ingen_tidning['Del av hushåll'] = np.where(df_ingen_tidning.duplicated(subset=['Familj','x','Folkbokföring - Gatuadress'], keep=False),"Ja","Nej")

    # Familjerepresentant (oldest in family)
    # See 'Fam.Admin'

    # Honoring "family" tagging - even if living on different locations
    #df_ingen_tidning = df_ingen_tidning.drop_duplicates(subset=['Familj','Folkbokföring - Gatuadress','Folkbokföring - Postnummer'], keep='first')
    
    # Working
    # df_ingen_tidning.drop_duplicates(subset=['Familj','x'], keep='last', inplace=True)

    #dfx = df_ingen_tidning.sort_values(['Fam.Admin'])
    #dfx = df_ingen_tidning.duplicated(subset=['Familj'], keep=False)
    #df_ingen_tidning['x'] = df_ingen_tidning.duplicated(subset=['Familj'], keep=False)

    # Add helper about 'Tidning' or not
    df_ingen_tidning['Tidning'] = [get_paper(famadmin, no_paper, in_familiy, in_household)
        for famadmin, no_paper, in_familiy, in_household 
        in zip (
            df_ingen_tidning['Fam.Admin'], 
            df_ingen_tidning['Ingen tidning'], 
            df_ingen_tidning['Del av familj'], 
            df_ingen_tidning['Del av hushåll'])]

    df_ingen_tidning['My Club Id'] = df_ingen_tidning['Övrig medlemsinfo'].apply(get_my_club_id)

    # Verification
    #df_test = df_ingen_tidning[df_ingen_tidning['Tidning'].str.contains("Nej")]
    #df_test = df_test.drop_duplicates(subset=['Familj','Fam.Admin','Tidning','Ingen tidning'], keep='first')
    #print(df_test)
    #exit(0)

    # Drop trick
    df_ingen_tidning.drop('x', 1, inplace=True)

    save_file("/usr/src/app/files/frisksport/frisksport_export_2.xlsx", 
        df_ingen_tidning[['IdrottsID','Personnummer','Förnamn', 'Efternamn',
            'Familj','Fam.Admin', 'Tidning', 'Ingen tidning', 
            'Del av familj', 'Del av hushåll',
            'Medlem sedan',
            'Medlem t.o.m.',
            'My Club Id',
            'Grupp',
            'Telefon mobil',
            'Telefon bostad',
            'Telefon arbete',
            'E-post privat',
            'E-post arbete',
            'E-post kontakt',
            'Folkbokföring - Gatuadress',
            'Folkbokföring - Postnummer',
            'Folkbokföring - Postort',
            'Kontaktadress - Gatuadress',
            'Kontaktadress - Postnummer',
            'Kontaktadress - Postort',
            'Kontaktadress - c/o adress',
            'Kontaktadress - Land',
            'Arbetsadress - c/o adress',
            'Medlemsnr.',
            'Övrig medlemsinfo']], True)
    #save_file("/usr/src/app/files/frisksport/exempel4.xlsx", df_ingen_tidning, True)
    if False:
        exit(0)

        print(df_ingen_tidning.head())
        save_file("/usr/src/app/files/frisksport/exempel3.xlsx", df_ingen_tidning, True)
        print("done")
        exit(0)

        # Exclude 'MC_IngenTidning'
        #df_ingen_tidning = df[df['Grupp'].notnull() &
        #    ~df['Grupp'].str.contains('MC_IngenTidning')]
        df_ingen_tidning = df_ingen_tidning[~df_ingen_tidning['Grupp'].str.contains('MC_IngenTidning')]
        #df = df.loc[~df_ingen_tidning]


        print(df)
        print(df_ingen_tidning[['Förnamn','Grupp','Personnummer','Folkbokföring - Gatuadress']])
        #print(df_ingen_tidning)
    print("Done")

print("Tidsåtgång: " + str(round((time.time() - start_time),1)) + " s")
print((" Klart (" + strftime("%Y-%m-%d %H:%M") + ") ").center(80, "-"))
