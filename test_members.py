# pylint: disable=import-error
import pandas as pd
import numpy as np
import os, sys
from pathlib import Path
from datetime import date
import time 
from time import strftime

'''
For testing thing regarding members  
'''
print("Testing members methods")

#def fill_key(key, ref):
#    if not key:
#        return ref
#    return key

#df = pd.DataFrame(np.arange(12).reshape(6, 2),
#                  columns=['Name', 'Val'])
#df['key'] = 
df = pd.DataFrame(columns=['key', 'ref', 'name', 'value'],
                  data=[
                    ['k1',None,'N1','A'],
                    [None,'k1','N2','B'],
                    [None,'k1','N3','C'],
                    ['k2',None,'N4','D'],
                    ['k3',None,'N5','E'],
                    [None,'k3','N6','F'],
                    [None,'k3','N7','G']])
print(df)
#df['key'] = df.apply(lambda x: fill_key(x['key'],x['ref']),axis=1)

print("groupby")
#df.groupby(["key","ref"], dropna=False).apply(print)
#df['parent_no'] = df.groupby(["key","ref"], dropna=False)['ref'].cumcount()+1 # Works
df['parent_no'] = df.groupby(["ref"], dropna=False)['ref'].cumcount()+1 # Also works
print(df)

goal = pd.DataFrame(columns=['key', 'ref', 'name', 'value', 'name1', 'value1', 'name2', 'value2'],
                  data=[
                    ['k1','k1','N1','A','N2','B','N3','C'],
                    ['k2',None,'N4','D'],
                    ['k3','k3','N5','E','N6','F','N7','G']])
goal.drop(columns=['key','ref'], inplace=True)
#print("goal")
#print(goal)

# Children
df_c = df[~df['ref'].notnull()].copy()
# Drop parent no
df_c.drop(columns=['parent_no'], inplace=True)
print("Children")
print(df_c)

# Parents
print("Parents")
df_p = df[df['ref'].notnull()]
print(df_p)

# Children merged with parents
#print("Merged")
#df_m = pd.merge(df_c, df_p, on='key')
#print(df_m)
#df_with_parents = pd.merge(df, df_parents, on='key')

#df1['row_number_by_group']=df1.groupby(['Product'])['Sales'].cumcount()+1

#.apply(print)

# Merge parent 1
print("merged")
#df_p1 = df_p[df_p['parent_no'] == 1]
#print(df_p1)
df_m = pd.merge(df_c, df_p[df_p['parent_no'] == 1], how='outer', left_on='key', right_on='ref', suffixes=('','1'))
df_m = pd.merge(df_m, df_p[df_p['parent_no'] == 2], how='outer', left_on='key', right_on='ref', suffixes=('','2'))
df_m.drop(columns=['key','ref','parent_no','key1','key2','ref1','ref2','parent_no2'], inplace=True)
print(df_m)

testing_isin = pd.DataFrame(columns=['name','value'],
                  data=[
                    ['a','a,OL A'],
                    ['b','b,OL A,OL D, OL B'],
                    ['c','c,OL A, OL B, OL C'],
                    ['d','d,OL B, OL C'],
                    ['e','e,OL A , OL B, OL C, OL D'],
                    ['','<tom>,OL A ,OL B, OL C'],
                    [None,'none,OL A ,OL B, OL C']])

def group_in_groups(groups, grp):
    if groups is np.nan:
        return False

    l = list(groups.split(",")) 
    l =[x.strip() for x in l]
    #print("Is '{}' in '{}'? {}".format(grp,groups, (grp in l)))
    return grp in l

print("".center(40,"-"))
#print(testing_isin['value'])
print(testing_isin['value'].apply(lambda x: group_in_groups(x,"OL B")))
print(testing_isin[testing_isin['value'].apply(lambda x: group_in_groups(x,"OL B"))])
print("".center(40,"-"))

for grp in ['OL A', 'OL B', 'OL D']:
    print("\n" + (" Testing: {} ".format(grp).center(40, "-")))
    #print(testing_isin[(testing_isin['name'].str.len()>0)]) # OK
    #print(testing_isin[(testing_isin['name'].str.len()==0)]) # Not OK, does not handle None
    #print(testing_isin[(testing_isin['name'].str.len()<1)]) # Not OK, does not handle None
    #print(testing_isin[~(testing_isin['name'].str.len()>0)]) # OK, hanterar tom sträng och None
    #print(testing_isin[(~(testing_isin['name'].str.len()>0))]) # OK, name får inte finnas angiven
    #print(testing_isin[(~(testing_isin['name'].str.len()>0)) & (testing_isin['value'].isin([grp]))])
    
    #print(testing_isin[((testing_isin['name'].str.len()>0))]) # OK, name måste finnas angiven
    print(testing_isin[(testing_isin['value'].apply(lambda x: group_in_groups(x,grp)))])

from openpyxl import Workbook
#from openpyxl.compat import range
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows

wb = Workbook()
wb = load_workbook(filename = 'files/contact-list/templates/template_youth_contactlist.xlsx')

# Kontaktlista
ws1 = wb.active

# testing_isin
print("to excel")
ws1["A10"] = "Hej"

print("loop items")
for rowidx, row in df.iterrows():
    #print(type(index))
    #print(index)
    #print('~~~~~~')

    #print(type(row))
    col = 1
    for c in row.values:
        print("rowindex: {}, col: {}, value {}".format(rowidx+2,col,c))
        ws1.cell(column=col, row=rowidx+2, value=c)
        col += 1
    #print('------')
    #ws3.cell(column=col, row=row, value="{0}".format(get_column_letter(col)))
    #ws1.cell(column=col, row=row, value=)
#for r in dataframe_to_rows(testing_isin, index=False, header=False):
    #print(r.index)
 #   for c in r:
 #       print("Value: {}".format(c))
    #ws1.append(r)
    #ws1[]

#for col, val in enumerate(mylist, start=1):
#    sheet.cell(row=2, column=col).value = val

dest_filename = 'files/contact-list/created/tmp/updated_book2.xlsx'
wb.save(filename = dest_filename)
